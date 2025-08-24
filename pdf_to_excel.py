#!/usr/bin/env python3
"""
PDF to Excel Converter
A comprehensive tool to convert PDF files to Excel format using multiple extraction methods.
"""

import os
import sys
import argparse
import logging
from pathlib import Path
from typing import List, Dict, Any, Optional
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

try:
    import pdfplumber
    import tabula
    import camelot
    import PyPDF2
except ImportError as e:
    print(f"Missing required dependency: {e}")
    print("Please install dependencies with: pip install -r requirements.txt")
    sys.exit(1)

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

class PDFToExcelConverter:
    """Main class for converting PDF files to Excel format."""
    
    def __init__(self, output_dir: str = "output"):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(exist_ok=True)
        
    def extract_text_with_pdfplumber(self, pdf_path: str) -> List[Dict[str, Any]]:
        """Extract text and tables using pdfplumber."""
        results = []
        
        try:
            with pdfplumber.open(pdf_path) as pdf:
                for page_num, page in enumerate(pdf.pages, 1):
                    page_data = {
                        'page': page_num,
                        'text': page.extract_text(),
                        'tables': []
                    }
                    
                    # Extract tables from the page
                    tables = page.extract_tables()
                    for table_num, table in enumerate(tables):
                        if table and any(any(cell for cell in row) for row in table):
                            # Convert table to DataFrame
                            df = pd.DataFrame(table[1:], columns=table[0])
                            page_data['tables'].append({
                                'table_num': table_num + 1,
                                'data': df
                            })
                    
                    results.append(page_data)
                    
        except Exception as e:
            logger.error(f"Error extracting with pdfplumber: {e}")
            
        return results
    
    def extract_tables_with_tabula(self, pdf_path: str) -> List[pd.DataFrame]:
        """Extract tables using tabula-py."""
        try:
            tables = tabula.read_pdf(pdf_path, pages='all', multiple_tables=True)
            return [df for df in tables if not df.empty]
        except Exception as e:
            logger.error(f"Error extracting with tabula: {e}")
            return []
    
    def extract_tables_with_camelot(self, pdf_path: str) -> List[pd.DataFrame]:
        """Extract tables using camelot-py."""
        try:
            tables = camelot.read_pdf(pdf_path, pages='all')
            return [table.df for table in tables if not table.df.empty]
        except Exception as e:
            logger.error(f"Error extracting with camelot: {e}")
            return []
    
    def extract_text_with_pypdf2(self, pdf_path: str) -> str:
        """Extract text using PyPDF2."""
        try:
            text = ""
            with open(pdf_path, 'rb') as file:
                pdf_reader = PyPDF2.PdfReader(file)
                for page in pdf_reader.pages:
                    text += page.extract_text() + "\n"
            return text
        except Exception as e:
            logger.error(f"Error extracting with PyPDF2: {e}")
            return ""
    
    def create_excel_workbook(self, data: List[Dict[str, Any]], 
                            tabula_tables: List[pd.DataFrame],
                            camelot_tables: List[pd.DataFrame],
                            pypdf2_text: str,
                            output_filename: str) -> str:
        """Create Excel workbook with extracted data."""
        output_path = self.output_dir / output_filename
        
        # Create workbook
        wb = openpyxl.Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Add pdfplumber data
        if data:
            ws_pdfplumber = wb.create_sheet("PDFPlumber_Extraction")
            row = 1
            
            for page_data in data:
                # Add page header
                ws_pdfplumber[f'A{row}'] = f"Page {page_data['page']}"
                ws_pdfplumber[f'A{row}'].font = Font(bold=True, size=14)
                ws_pdfplumber[f'A{row}'].fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                row += 1
                
                # Add text content
                if page_data['text']:
                    ws_pdfplumber[f'A{row}'] = "Text Content:"
                    ws_pdfplumber[f'A{row}'].font = Font(bold=True)
                    row += 1
                    
                    # Split text into lines and add to Excel
                    text_lines = page_data['text'].split('\n')
                    for line in text_lines[:100]:  # Limit to first 100 lines to avoid Excel row limit
                        if line.strip():
                            ws_pdfplumber[f'A{row}'] = line.strip()
                            row += 1
                    
                    row += 2
                
                # Add tables
                for table_info in page_data['tables']:
                    ws_pdfplumber[f'A{row}'] = f"Table {table_info['table_num']}:"
                    ws_pdfplumber[f'A{row}'].font = Font(bold=True)
                    row += 1
                    
                    # Add table data
                    for r in dataframe_to_rows(table_info['data'], index=False, header=True):
                        for col, value in enumerate(r, 1):
                            ws_pdfplumber.cell(row=row, column=col, value=value)
                        row += 1
                    
                    row += 2
        
        # Add tabula tables
        if tabula_tables:
            ws_tabula = wb.create_sheet("Tabula_Tables")
            row = 1
            
            for i, table in enumerate(tabula_tables, 1):
                ws_tabula[f'A{row}'] = f"Table {i} (Tabula):"
                ws_tabula[f'A{row}'].font = Font(bold=True)
                row += 1
                
                for r in dataframe_to_rows(table, index=False, header=True):
                    for col, value in enumerate(r, 1):
                        ws_tabula.cell(row=row, column=col, value=value)
                    row += 1
                
                row += 2
        
        # Add camelot tables
        if camelot_tables:
            ws_camelot = wb.create_sheet("Camelot_Tables")
            row = 1
            
            for i, table in enumerate(camelot_tables, 1):
                ws_camelot[f'A{row}'] = f"Table {i} (Camelot):"
                ws_camelot[f'A{row}'].font = Font(bold=True)
                row += 1
                
                for r in dataframe_to_rows(table, index=False, header=True):
                    for col, value in enumerate(r, 1):
                        ws_camelot.cell(row=row, column=col, value=value)
                    row += 1
                
                row += 2
        
        # Add PyPDF2 text
        if pypdf2_text:
            ws_pypdf2 = wb.create_sheet("PyPDF2_Text")
            row = 1
            
            ws_pypdf2[f'A{row}'] = "Extracted Text (PyPDF2):"
            ws_pypdf2[f'A{row}'].font = Font(bold=True)
            row += 1
            
            # Split text into lines and add to Excel
            text_lines = pypdf2_text.split('\n')
            for line in text_lines[:100]:  # Limit to first 100 lines
                if line.strip():
                    ws_pdfplumber[f'A{row}'] = line.strip()
                    row += 1
        
        # Auto-adjust column widths
        for ws in wb.worksheets:
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save workbook
        wb.save(output_path)
        return str(output_path)
    
    def convert_pdf_to_excel(self, pdf_path: str, output_filename: Optional[str] = None) -> str:
        """Main conversion method."""
        pdf_path = Path(pdf_path)
        
        if not pdf_path.exists():
            raise FileNotFoundError(f"PDF file not found: {pdf_path}")
        
        if output_filename is None:
            output_filename = f"{pdf_path.stem}_converted.xlsx"
        
        logger.info(f"Converting {pdf_path} to Excel...")
        
        # Extract data using different methods
        logger.info("Extracting data with pdfplumber...")
        pdfplumber_data = self.extract_text_with_pdfplumber(str(pdf_path))
        
        logger.info("Extracting tables with tabula...")
        tabula_tables = self.extract_tables_with_tabula(str(pdf_path))
        
        logger.info("Extracting tables with camelot...")
        camelot_tables = self.extract_tables_with_camelot(str(pdf_path))
        
        logger.info("Extracting text with PyPDF2...")
        pypdf2_text = self.extract_text_with_pypdf2(str(pdf_path))
        
        # Create Excel workbook
        logger.info("Creating Excel workbook...")
        output_path = self.create_excel_workbook(
            pdfplumber_data, tabula_tables, camelot_tables, pypdf2_text, output_filename
        )
        
        logger.info(f"Conversion completed! Output saved to: {output_path}")
        return output_path

def main():
    """Command-line interface."""
    parser = argparse.ArgumentParser(description="Convert PDF files to Excel format")
    parser.add_argument("pdf_file", help="Path to the PDF file to convert")
    parser.add_argument("-o", "--output", help="Output Excel filename (optional)")
    parser.add_argument("-d", "--output-dir", default="output", help="Output directory (default: output)")
    
    args = parser.parse_args()
    
    try:
        converter = PDFToExcelConverter(args.output_dir)
        output_path = converter.convert_pdf_to_excel(args.pdf_file, args.output)
        print(f"\n✅ Conversion successful!")
        print(f"📁 Output file: {output_path}")
        
    except Exception as e:
        logger.error(f"Conversion failed: {e}")
        sys.exit(1)

if __name__ == "__main__":
    main()